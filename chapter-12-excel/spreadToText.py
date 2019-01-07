#!python3
# The program should open a spreadsheet and write the cells of column A into one text file, the cells of column B into another text file, and so on.

import os
import sys
import openpyxl

try:
    assert len(sys.argv) == 2
    assert sys.argv[1].endswith('.xlsx')
    assert (os.path.exists(sys.argv[1]))
except AssertionError:
    print("Command line argument is not valid. Enter a valid excel file")
    quit()

excelfile = sys.argv[1]

wb = openpyxl.load_workbook(excelfile)
ws = wb.active 

print("Starting...")
for col in range(1, ws.max_column + 1):
    print("Starting to transfer column " + str(col))
    fileObj = open(excelfile + "_text_of_col" + str(col) + ".txt", "w")
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=col).value 
        fileObj.write(str(value) + "\n")
    fileObj.close()    
    print("Done transferring column " + str(col))    

print("Completed.")
