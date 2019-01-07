#!python 3 
# multiplicationTable.py takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet.

import os, sys, openpyxl 
from openpyxl.styles import Font

assert (len(sys.argv) == 2 and type(int(sys.argv[1])) == int)
N = int(sys.argv[1])
wb = openpyxl.Workbook()
ws = wb.active 

# Adding 1st row and 1st col 
boldFont = Font(bold=True, italic=True)
print("Adding headers...")
for i in range(1, N+1):
    # Adding 1st row header and styling it 
    row_i = i + 1 
    cell_row = ws.cell(row=1, column=row_i)
    cell_row.font = boldFont
    cell_row.value = i 
    # Adding 1st column header and styling it 
    col_i = i + 1 
    cell_col = ws.cell(row=col_i, column=1)
    cell_col.font = boldFont
    cell_col.value = i

print("Headers added. Adding multiples...")
# Adding multiples
for row_number in range(2, N+2):
    row_multiple = row_number - 1
    for col_number in range(2, N+2):
        col_multiple = col_number - 1
        ws.cell(row=row_number, column=col_number, value=row_multiple*col_multiple) 
print("Multiples added... Saving file...")

wb.save('multiplicationTable_{}x{}.xlsx'.format(N, N))
print("File saved.")
